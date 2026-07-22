"""
excel_normalizer.py — 把「任意來源 Excel」轉成標準每日配送表

標準格式 (5 欄)：車號 / 店家名稱 / 店家地址 / 瓶數 / 品項
來源可能：
  - 欄位名不同 (車輛/店名/地址/數量/品名...) → 寬鬆對應
  - 沒有「車號」欄 → 留空，由 caller 決定怎麼分車
  - 沒有「品項」欄 → 留空（純鮮奶配送）
  - 地址欄可能夾雜「-隨貨附發票」之類後綴 → 保留原值(地理編碼時再清理)
  - 多餘列/空列 → 跳過

產出：DataFame(標準5欄) + 存檔 每日配送_YYYYMMDD.xlsx 到 out_dir
"""

import os
import re
import datetime

try:
    import openpyxl
except ImportError:
    openpyxl = None

# 欄位別名對應（小寫去空白後比對）
_VEH = ["車號", "車輛", "路線編號", "路線", "route", "vehicle", "車", "車牌"]
_NAME = ["店家名稱", "名稱", "店名", "客戶名稱", "客戶簡稱", "客戶", "簡稱", "店家", "name", "店", "對象"]
_ADDR = ["店家地址", "送貨地址", "地址", "客戶地址", "送貨", "address", "addr", "位置"]
_QTY = ["瓶數", "數量", "箱數", "瓶量", "qty", "bottles", "count", "件數", "瓶", "量"]
_ITEM = ["品項", "品名", "貨品", "項目", "item", "items"]
_FUEL = ["油資單價", "油資", "油錢單價", "fuel", "fuel_cost", "fuel_cost_per_km", "元每km"]
_CONS = ["特殊需求", "特殊要求", "需求", "備註", "備註說明", "constraint", "note", "remark"]

# 台灣縣市關鍵字（用於無車號時的地理分車）
_REGION_KEYWORDS = [
    ("台北", "台北"), ("新北", "新北"), ("基隆", "基隆"), ("桃園", "桃園"),
    ("新竹", "新竹"), ("苗栗", "苗栗"), ("台中", "台中"), ("彰化", "彰化"),
    ("南投", "南投"), ("雲林", "雲林"), ("嘉義", "嘉義"), ("台南", "台南"),
    ("高雄", "高雄"), ("屏東", "屏東"), ("宜蘭", "宜蘭"), ("花蓮", "花蓮"),
    ("台東", "台東"), ("澎湖", "澎湖"),
]


# 鮮奶關鍵字 (含「鮮乳/鮮奶/牛奶」才計瓶數; 保久乳/糖漿/果泥等歸品項)
_FRESH_MILK_HINTS = ["鮮乳", "鮮奶", "牛奶"]


def _norm_addr(addr):
    """去後綴 + 統一全半形/空白/括號/結尾標點，供同店比對。"""
    a = (addr or "").strip()
    a = a.replace("臺", "台").replace("　", " ").replace("(", "").replace(")", "")
    a = re.sub(r"\s+", "", a)
    a = a.strip(".,-_/\\ ")
    return a


def _is_milk(name):
    return any(h in (name or "") for h in _FRESH_MILK_HINTS)


def _norm(h):
    return (h or "").strip().lower().replace(" ", "").replace("　", "")


def _match_col(headers_norm, aliases):
    """在表頭們(headers_norm: {norm表頭: 原表頭}) 中, 找包含任一別名的欄位。
    別名可能是表頭的子串(如 別名'地址' 命中表頭'送貨地址')。"""
    # 先嘗試「完全相等」(最快)
    for a in aliases:
        if _norm(a) in headers_norm:
            return headers_norm[_norm(a)]
    # 再嘗試「別名是表頭的子串」(寬鬆)
    for h_norm, h_orig in headers_norm.items():
        for a in aliases:
            if _norm(a) and _norm(a) in h_norm:
                return h_orig
    return None


def _clean_int(v):
    try:
        if v is None or v == "":
            return 0
        if isinstance(v, (int, float)):
            return int(v)
        # 去掉逗號/空白後取數字 (保留負號, 銷貨退回為負)
        m = re.search(r"-?\d+", str(v))
        return int(m.group()) if m else 0
    except Exception:
        return 0


def normalize_excel(path, out_dir, default_vehicle="車01", date_str=None):
    """
    讀任意來源 Excel，轉成標準 4 欄 DataFrame。
    rows: [(車號, 店家名稱, 店家地址, 瓶數, 品項), ...]  (車號可能為 "")
      skipped: [(原始店名或標記, 原因), ...]
      out_path: 存的檔路徑 (每日配送_YYYYMMDD.xlsx) 或 None(若沒裝 openpyxl)
    """
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xls":
        # 舊版 Excel 97-2003 → 用 xlrd 讀
        try:
            import xlrd
        except ImportError:
            raise RuntimeError("讀 .xls 需要 xlrd: uv pip install xlrd==2.0.1")
        book = xlrd.open_workbook(path)
        sh = book.sheet_by_index(0)
        raw = []
        for row_idx in range(sh.nrows):
            raw.append([sh.cell_value(row_idx, c) for c in range(sh.ncols)])
    else:
        # .xlsx / .xlsm → openpyxl
        if openpyxl is None:
            raise RuntimeError("請先安裝 openpyxl: uv pip install openpyxl")
        wb = openpyxl.load_workbook(path, data_only=True)
        # 優先讀「工作表2」(實際資料表)；找不到再回退 active (與 sales_to_dispatch 一致，
        # 否則 wb.active 可能指向空白模板表，導致有資料的列被誤判「店家與地址皆空」而跳過)
        if "工作表2" in wb.sheetnames:
            ws = wb["工作表2"]
        else:
            ws = wb.active
        raw = list(ws.iter_rows(values_only=True))
    if not raw:
        return [], [], None

    # 自動偵測表頭行：掃前 15 列，找第一列出現已知欄位別名的列
    header_row_idx = 0
    for ridx, r in enumerate(raw[:15]):
        cells = [_norm(str(c)) for c in r if c is not None and str(c).strip() != ""]
        if not cells:
            continue
        # 只要這列含 店家/地址/瓶數 任一別名，就當表頭
        hit = any(any(_norm(a) in c for c in cells) for a in (_NAME + _ADDR + _QTY))
        if hit:
            header_row_idx = ridx
            break
    headers = [str(h) if h is not None else "" for h in raw[header_row_idx]]
    hnorm = {_norm(h): h for h in headers}

    col_name = _match_col(hnorm, _NAME)
    col_addr = _match_col(hnorm, _ADDR)
    col_qty = _match_col(hnorm, _QTY)
    col_veh = _match_col(hnorm, _VEH)
    col_fuel = _match_col(hnorm, _FUEL)
    col_item = _match_col(hnorm, _ITEM)
    col_cons = _match_col(hnorm, _CONS)

    skipped = []
    rows = []
    fuel_cost = None   # 全車共用油資單價（取第一個有效值）
    _last = {"name": "", "addr": ""}  # 續行繼承用: 記上一筆有值的店名/地址
    _shops = {}  # 同店合併: akey -> 合併後店資料
    for i, r in enumerate(raw[header_row_idx + 1:], header_row_idx + 2):
        if all(v is None or v == "" for v in r):
            continue  # 空列跳過
        # 用表頭欄位名去定位索引（找不到該欄就跳過該欄）
        def _val(col):
            if not col:
                return ""
            try:
                idx = headers.index(col)
                v = r[idx] if idx < len(r) else ""
                return "" if v is None else v
            except (ValueError, IndexError):
                return ""
        name = str(_val(col_name)).strip()
        addr = str(_val(col_addr)).strip()
        qty = _clean_int(_val(col_qty))
        veh = str(_val(col_veh)).strip()
        item = str(_val(col_item)).strip()
        cons = str(_val(col_cons)).strip()
        # 續行繼承 (Bug: 同列前面空白 = 承襲上一列店家/地址)
        #   客戶簡稱 與 送貨地址 皆空 → 繼承上一筆
        #   僅地址空白(店名重複) → 繼承上一筆地址
        #   僅店名空白(地址有值) → 繼承上一筆店名
        if not name and not addr:
            name = _last.get("name", "")
            addr = _last.get("addr", "")
        elif not addr and name:
            addr = _last.get("addr", "")
        elif not name and addr:
            name = _last.get("name", "")
        if name:
            _last["name"] = name
        if addr:
            _last["addr"] = addr
        if fuel_cost is None:   # 只取第一個有效油資值
            fv = _val(col_fuel)
            if fv not in (None, ""):
                try:
                    fv = float(fv)
                    if fv > 0:
                        fuel_cost = fv
                except (ValueError, TypeError):
                    pass
        if not name and not addr:
            skipped.append((f"第{i}列", "店家與地址皆空"))
            continue
        if not addr:
            skipped.append((name or f"第{i}列", "缺少店家地址"))
            continue
        # 同店合併 (按正規化地址): 避免續行/重複地址被拆成多筆, JOJO 重複排同一地址
        akey = _norm_addr(addr)
        if akey not in _shops:
            _shops[akey] = {"veh": veh, "name": name or f"店家{i}", "addr": addr,
                            "milk": 0.0, "items": {}, "cons": cons}
        sh = _shops[akey]
        if veh and not sh["veh"]:
            sh["veh"] = veh
        if cons:
            sh["cons"] = cons
        if name and sh["name"].startswith("店家"):
            sh["name"] = name
        if _is_milk(item):
            sh["milk"] += qty
        elif item:
            if item in sh["items"]:
                sh["items"][item] += qty
            else:
                sh["items"][item] = qty

    # 合併後輸出 (與 sales_to_dispatch 一致): 瓶數(鮮奶) + 品項欄(非鮮奶彙總)
    for akey, sh in _shops.items():
        milk_int = int(round(sh["milk"]))
        item_str = ", ".join(f"{nm}{qty:.0f}" for nm, qty in sh["items"].items())
        # 退貨相抵後無貨 (milk<=0 且無其他品項) → 取消訂單, 不排入路線
        if milk_int <= 0 and not item_str:
            continue
        # 有品項但鮮奶相抵成負/0 → 瓶數歸 0 (仍有非鮮奶貨要送)
        out_milk = milk_int if milk_int > 0 else 0
        cons_val = sh.get("cons", "")
        rows.append((sh["veh"], sh["name"], sh["addr"], out_milk, item_str, cons_val))

    if not rows:
        return [], skipped, None, fuel_cost

    # 存成標準檔：每日配送_YYYYMMDD.xlsx (5欄)
    date_str = date_str or datetime.datetime.now().strftime("%Y%m%d")
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, f"每日配送_{date_str}.xlsx")
    from openpyxl import Workbook
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "每日配送"
    ws2.append(["車號", "店家名稱", "店家地址", "瓶數", "品項", "特殊需求"])
    for veh, name, addr, qty, item, cons in rows:
        ws2.append([veh, name, addr, qty, item, cons])
    wb2.save(out_path)
    return rows, skipped, out_path, fuel_cost


def region_of(addr: str) -> str:
    """從地址判斷縣市(用於無車號時的地理分車)。"""
    for kw, label in _REGION_KEYWORDS:
        if kw in addr:
            return label
    return "其他"


def auto_assign_vehicles(rows):
    """
    來源無車號時，由 Agent 自己決定車號。
    策略：先全部併為 1 台車（車01）；
          若 caller 發現 1 台車跑完超過 17:00 回倉，
          再呼叫 split_by_region 按地理區域拆成多台。
    這裡回傳「全併 1 台車」版本（最保守起點），
    拆車由 plan 層回傳超時資訊後再呼叫 split_by_region。
    """
    return [("車01", n, a, q) for (_, n, a, q) in rows]


_REGION_LABEL_TO_VEH = {}
def split_by_region(rows):
    """把無車號 rows 按縣市拆成多台車：每台車 = 一個縣市。"""
    groups = {}
    order = []
    for (_, n, a, q) in rows:
        reg = region_of(a)
        if reg not in groups:
            groups[reg] = []
            order.append(reg)
        groups[reg].append((n, a, q))
    out = []
    for i, reg in enumerate(order, 1):
        veh = f"車{i:02d}"
        for (n, a, q) in groups[reg]:
            out.append((veh, n, a, q))
    return out


if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1:
        p = sys.argv[1]
        out = sys.argv[2] if len(sys.argv) > 2 else os.path.dirname(p)
        rows, skipped, outp, fuel = normalize_excel(p, out)
        print(f"讀到 {len(rows)} 筆，跳過 {len(skipped)} 筆，油資單價={fuel}")
        for r in rows[:5]:
            print(" ", r)
        if outp:
            print("標準檔已存:", outp)
