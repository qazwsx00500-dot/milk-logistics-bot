"""
ann_archive.py — ANN 的「本機存檔」工具

職責（來自 SOUL.md / TEAM.md 的職權劃分）：
  JOJO 負責跑路線規劃、產出路線圖＋報表、在 LINE 回傳結果給傑夫；
  路線圖/報表產出後，JOJO 把「複本」交給 ANN（即 Render 雲端的
  /report、/dispatch、/route_map、/workbook 端點），由 ANN 負責把這些
  成果「本機存檔」到 OneDrive 桌面對應資料夾，並產出整合 Excel。

  JOJO 跑完即停止、等待下一次分配；本機存檔不是 JOJO 的活，是 ANN 的活。

做法：
  直接複用既有 sync_from_render.py 的抓取邏輯（Render 雲端 → 本機 OneDrive 桌面），
  再額外加一步：把當日 route_report.csv + dispatch.csv 整合成「整合報表.xlsx」
  （3 分頁：路線總表 / 各車派車單 / 總計）。整合 Excel 由 ANN 產出，JOJO 不再產。

用法：
  python ann_archive.py                 # 抓今天，存 OneDrive 桌面 + 產整合 Excel
  python ann_archive.py --date 2026-07-19
  python ann_archive.py --base https://milk-logistics-bot.onrender.com

依賴：標準庫 + logistics_agent.py（僅取 REPORT_DIR / DISPATCH_DIR 路徑）。
"""

import argparse
import os
import ssl
import sys
import csv
import urllib.error
import urllib.request
from datetime import datetime, timedelta, timezone

import logistics_agent as L

BASE = os.environ.get("RENDER_BASE_URL", "https://milk-logistics-bot.onrender.com")

# (端點, 目標資料夾屬性, 檔名) —— 與 sync_from_render.py 一致
JOBS = [
    ("/report",      "REPORT_DIR",  "route_report.html"),
    ("/report.csv",  "REPORT_DIR",  "route_report.csv"),
    ("/route_map",   "REPORT_DIR",  "route_map.html"),
    ("/dispatch",    "DISPATCH_DIR", "dispatch.html"),
    ("/dispatch.csv", "DISPATCH_DIR", "dispatch.csv"),
    ("/workbook",    "DISPATCH_DIR", "整合報表.xlsx"),
    ("/map_png",     "DISPATCH_DIR", "route_map.png"),
]

_EMPTY_HINTS = ("尚未產生報表", "尚未產生派車單", "尚無路線地圖")


def _fetch(url, timeout=40):
    ctx = ssl.create_default_context()
    req = urllib.request.Request(url, headers={"User-Agent": "ann-archive/1.0"})
    with urllib.request.urlopen(req, timeout=timeout, context=ctx) as r:
        return r.status, r.read()


def _today_tw():
    return (datetime.now(timezone.utc) + timedelta(hours=8)).strftime("%Y-%m-%d")


def pull(day, base):
    """從雲端抓 JOJO 產出的複本，存到本機 OneDrive 桌面對應資料夾。回傳 (ok, empty, fail)。"""
    base = (base or BASE).rstrip("/")
    dirs = {"REPORT_DIR": L.REPORT_DIR, "DISPATCH_DIR": L.DISPATCH_DIR}
    print(f"🔄 ANN 從 {base} 抓 {day} 的 JOJO 成果，本機存檔 …")
    ok = empty = fail = 0
    for path, dir_attr, fname in JOBS:
        url = base + path
        try:
            status, body = _fetch(url)
        except urllib.error.HTTPError as e:
            print(f"  ✗ {path} → HTTP {e.code}")
            fail += 1
            continue
        except Exception as e:
            print(f"  ✗ {path} → {type(e).__name__}: {str(e)[:60]}")
            fail += 1
            continue
        head = body[:80].decode("utf-8", errors="ignore")
        if any(h in head for h in _EMPTY_HINTS):
            print(f"  ⚠ {path} → 雲端尚未產生（跳過）")
            empty += 1
            continue
        day_dir = os.path.join(dirs[dir_attr], day)
        os.makedirs(day_dir, exist_ok=True)
        out = os.path.join(day_dir, fname)
        with open(out, "wb") as f:
            f.write(body)
        print(f"  ✓ {path} → {out}  ({len(body):,} bytes)")
        ok += 1
    return ok, empty, fail


def build_integrated_excel(day):
    """把當日 route_report.csv + dispatch.csv 整合成『整合報表_<日期>.xlsx』。
    分頁：①路線總表 ②各車派車單 ③總計。無資料則回 None。"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("  ⚠ 缺 openpyxl（uv pip install openpyxl），跳過整合 Excel。")
        return None

    report_csv = os.path.join(L.REPORT_DIR, day, "route_report.csv")
    dispatch_csv = os.path.join(L.DISPATCH_DIR, day, "dispatch.csv")
    if not os.path.exists(report_csv):
        print(f"  ⚠ 找不到 {report_csv}，跳過整合 Excel。")
        return None

    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="0B6B3A")
    veh_font = Font(bold=True, color="FFFFFF")
    veh_fill = PatternFill("solid", fgColor="C0392B")
    center = Alignment(horizontal="center")

    def _style_header(ws, row=1):
        for c in ws[row]:
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = center

    def _read_csv(p):
        with open(p, encoding="utf-8-sig", newline="") as f:
            return list(csv.reader(f))

    wb = openpyxl.Workbook()
    # ① 路線總表
    ws1 = wb.active
    ws1.title = "路線總表"
    rows = _read_csv(report_csv)
    for r in rows:
        ws1.append(r)
    if ws1.max_row:
        _style_header(ws1)
        ws1.freeze_panes = "A2"

    # ② 各車派車單
    ws2 = wb.create_sheet("各車派車單")
    if os.path.exists(dispatch_csv):
        for r in _read_csv(dispatch_csv):
            ws2.append(r)
        if ws2.max_row:
            _style_header(ws2)
            ws2.freeze_panes = "A2"

    # ③ 總計（簡單彙總）
    ws3 = wb.create_sheet("總計")
    ws3.append(["項目", "數值"])
    ws3.append(["出車數", len({r[0] for r in rows[1:] if r and r[0]})])
    ws3.append(["配送店數", max(0, len(rows) - 1)])
    _style_header(ws3)

    out_path = os.path.join(L.DISPATCH_DIR, day, f"整合報表_{day}.xlsx")
    wb.save(out_path)
    print(f"  ✓ 整合報表：{out_path}")
    return out_path


def main():
    ap = argparse.ArgumentParser(description="ANN：把 JOJO 的雲端成果本機存檔 + 產整合 Excel")
    ap.add_argument("--date", help="資料夾日期(YYYY-MM-DD)，預設今天(台灣時間)")
    ap.add_argument("--base", help=f"Render 網址，預設 {BASE}")
    ap.add_argument("--no-excel", action="store_true", help="只抓檔、不產整合 Excel")
    args = ap.parse_args()

    day = args.date or _today_tw()
    ok, empty, fail = pull(day, args.base)
    if not args.no_excel:
        build_integrated_excel(day)

    print(f"\n完成：成功 {ok}、尚未產生 {empty}、失敗 {fail}")
    if ok:
        print(f"📁 報表(本機存檔)：{os.path.join(L.REPORT_DIR, day)}")
        print(f"🚚 派車單(本機存檔)：{os.path.join(L.DISPATCH_DIR, day)}")
    if empty and not ok:
        print("（雲端這天還沒跑過規劃：先在 LINE 傳一次 Excel 讓 JOJO 跑，再執行本程式）")
    return 0 if fail == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
