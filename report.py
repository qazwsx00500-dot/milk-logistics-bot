"""
report.py — 路線報表產出

產出：
  - route_report.html : 可列印/手機檢視的路線排程報表 (含每站 ETA、瓶數、下貨時間)
  - route_report.csv  : 給司機/後台用的逐站 CSV

報表欄位（每站）：序號 | 店家 | 地址 | 瓶數 | 下貨(秒) | 預計到達 | 預計離開 | 累計里程
"""

import csv

# ---- PNG 下載按鈕（瀏覽器端 html2canvas 截圖，雲端本機皆可用） ----
_PNG_BTN = ('<button class="dlbtn" style="display:inline-block;background:#0b6b3a;color:#fff;'
            'border:none;border-radius:8px;padding:10px 16px;font-size:14px;font-weight:700;'
            'cursor:pointer;margin:4px 0 12px;box-shadow:0 1px 4px rgba(0,0,0,.2)" '
            'onclick="dlReportPNG()">\U0001f4e5 \u4e0b\u8f09\u5831\u8868 PNG</button>')

_PNG_SCRIPT = ('<script src="https://cdn.jsdelivr.net/npm/html2canvas@1.4.1/dist/html2canvas.min.js"></script>'
  '<script>'
  'function dlReportPNG(){'
  '  var btn=document.querySelector(".dlbtn"); var old=btn.textContent;'
  '  btn.textContent="\u7522\u751f\u4e2d\u2026"; btn.disabled=true;'
  '  html2canvas(document.querySelector(".wrap"),{scale:2,backgroundColor:"#f4f5f7"}).then(function(c){'
  '    var a=document.createElement("a");'
  '    a.download="\u914d\u9001\u5831\u8868_"+new Date().toISOString().slice(0,10)+".png";'
  '    a.href=c.toDataURL("image/png"); a.click();'
  '    btn.textContent=old; btn.disabled=false;'
  '  }).catch(function(e){alert("\u7522\u751f PNG \u5931\u6557:"+e);btn.textContent=old;btn.disabled=false;});'
  '}'
  '</script>')

import json
import os
from datetime import datetime, timedelta


def _hhmm(hour_float):
    """24h 小數 -> HH:MM。"""
    h = int(hour_float) % 24
    m = int(round((hour_float - int(hour_float)) * 60))
    if m == 60:
        h += 1; m = 0
    return f"{h:02d}:{m:02d}"


def _items_str(stop):
    """回傳該站非鮮奶品項的顯示字串，如 '冰勃朗1(箱),鳳梨果泥3(包)'；無則空字。"""
    items = getattr(stop, "items", None) or {}
    if not items:
        return ""
    return ", ".join(f"{n}{v['qty']:.0f}({v['unit']})" for n, v in items.items())


def _eta_rows(result, depot):
    """把所有路線攤平成逐站報表列。"""
    rows = []
    for ri, rt in enumerate(result.routes, 1):
        v = rt["vehicle"]
        seq = [depot] + rt["stops"] + [depot]
        # 累計里程
        cum_km = 0.0
        prev = depot
        for si, s in enumerate(rt["stops"]):
            # 行車里程 (用該段直線? 這裡用 result 的 distance 反推不精確，改用點對直線估算僅供參考)
            arrive, leave = rt["etas"][si]
            qty = getattr(s, "demand", 0)
            svc = int(round(getattr(s, "service_time", 0) or 0))
            rows.append({
                "route_no": ri,
                "vehicle": v.name,
                "seq": si + 1,
                "name": s.name,
                "address": getattr(s, "address", ""),
                "qty": int(qty) if qty == int(qty) else qty,
                "items": _items_str(s),
                "service_sec": svc,
                "arrive": _hhmm(arrive),
                "leave": _hhmm(leave),
            })
    return rows


def build_html(result, depot, out_path, meta=None):
    rows = _eta_rows(result, depot)
    gen = datetime.now().strftime("%Y-%m-%d %H:%M")
    src_map = {"haversine": "直線估算", "osrm": "OSRM 真實道路",
               "fallback": "直線估算(OSRM/Google 失敗降級)", "google": "Google Maps 真實道路"}
    src = src_map.get(result.distance_source, result.distance_source)
    meta = meta or {}

    # 路線摘要卡
    route_cards = ""
    for ri, rt in enumerate(result.routes, 1):
        v = rt["vehicle"]
        end = _hhmm(rt.get("end_hour", 0))
        stops_html = ""
        for si, s in enumerate(rt["stops"]):
            a, lv = rt["etas"][si]
            qty = getattr(s, "qty", s.demand)
            item_txt = _items_str(s)
            stops_html += (
                f"<tr><td>{si+1}</td><td><b>{s.name}</b><br><span class='addr'>{getattr(s,'address','')}</span></td>"
                f"<td class='num'>{int(qty) if qty==int(qty) else qty}</td>"
                f"<td class='num'>{item_txt or '—'}</td>"
                f"<td class='num'>{_hhmm(a)}</td><td class='num'>{_hhmm(lv)}</td></tr>"
            )
        route_cards += f"""
        <div class="card">
          <div class="card-h">路線 {ri} · {v.name}</div>
          <div class="card-meta">
            站數 {len(rt['stops'])} ｜ 里程 {rt['distance_km']:.1f} km ｜
            載重 {rt['load']:.0f} ｜ 回到倉庫 {end}
            {'' if rt['feasible'] else ' <span class="warn">⚠ 超限制</span>'}
          </div>
          <table>
            <thead><tr><th>#</th><th>店家 / 地址</th><th>瓶數</th><th>品項</th><th>到達</th><th>離開</th></tr></thead>
            <tbody>{stops_html}</tbody>
          </table>
        </div>"""

    skipped = getattr(result, "skipped", [])
    skip_html = ""
    if skipped:
        items = "".join(f"<li>{name} — {reason}</li>" for name, reason in skipped)
        skip_html = f"<div class='skip'><b>⚠ 未排入 {len(skipped)} 家（地址無法定位或超車隊容量）：</b><ul>{items}</ul></div>"

    html = f"""<!DOCTYPE html>
<html lang="zh-TW"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>鮮奶配送路線報表</title>
<style>
 body{{font-family:-apple-system,"Microsoft JhengHei",sans-serif;margin:0;background:#f4f5f7;color:#222;}}
 .wrap{{max-width:960px;margin:0 auto;padding:18px;}}
 header{{background:#0b6b3a;color:#fff;padding:16px 18px;border-radius:10px;}}
 header h1{{margin:0;font-size:20px;}}
 header .sub{{opacity:.9;font-size:13px;margin-top:4px;}}
 .summary{{display:flex;gap:10px;flex-wrap:wrap;margin:14px 0;}}
 .kpi{{background:#fff;border-radius:10px;padding:10px 14px;flex:1;min-width:120px;box-shadow:0 1px 3px rgba(0,0,0,.08);}}
 .kpi .v{{font-size:20px;font-weight:700;}}
 .kpi .l{{font-size:12px;color:#666;}}
 .card{{background:#fff;border-radius:10px;padding:12px 14px;margin-bottom:12px;box-shadow:0 1px 3px rgba(0,0,0,.08);}}
 .card-h{{font-weight:700;font-size:15px;margin-bottom:4px;}}
 .card-meta{{font-size:12px;color:#555;margin-bottom:8px;}}
 table{{width:100%;border-collapse:collapse;font-size:13px;}}
 th,td{{text-align:left;padding:6px 6px;border-bottom:1px solid #eee;}}
 th{{color:#666;font-weight:600;}}
 td.num,th.num{{text-align:right;font-variant-numeric:tabular-nums;}}
 .addr{{color:#888;font-size:11px;}}
 .warn{{color:#c0392b;font-weight:700;}}
 .ok{{color:#0b6b3a;font-weight:700;}}
 .skip{{background:#fff3cd;border:1px solid #ffe69c;border-radius:10px;padding:10px 14px;font-size:13px;}}
 .skip ul{{margin:6px 0 0;padding-left:18px;}}
 footer{{text-align:center;color:#999;font-size:12px;margin:16px 0;}}
</style></head>
<body><div class="wrap">
<header><h1>🥛 鮮奶配送路線報表</h1>
<div class="sub">出發 {_hhmm(meta.get('start_hour',8.0))} ｜ 距離來源 {src} ｜ 生成 {gen}</div></header>
<div class="summary">
  <div class="kpi"><div class="v">{len(result.routes)}</div><div class="l">出車數</div></div>
  <div class="kpi"><div class="v">{sum(len(r['stops']) for r in result.routes)}</div><div class="l">配送店數</div></div>
  <div class="kpi"><div class="v">{result.total_distance_km:.0f}</div><div class="l">總里程 km</div></div>
  <div class="kpi"><div class="v">{result.total_load:.0f}</div><div class="l">總瓶數</div></div>
</div>
{route_cards}
{skip_html}
<footer>本報表由物流路線規劃 Agent 雛形產出 · 時間含每瓶 15 秒下貨</footer>
</div></body></html>"""
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(html)
    return out_path


def build_csv(result, depot, out_path):
    rows = _eta_rows(result, depot)
    with open(out_path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(["路線", "車輛", "序號", "店家", "地址", "瓶數", "品項", "下貨秒數", "預計到達", "預計離開"])
        for r in rows:
            w.writerow([r["route_no"], r["vehicle"], r["seq"], r["name"], r["address"],
                        r["qty"], r["items"], r["service_sec"], r["arrive"], r["leave"]])
    return out_path


# ---------- 分組版（照車號） ----------

def _hhmm(hour_float):
    h = int(hour_float) % 24
    m = int(round((hour_float - int(hour_float)) * 60))
    if m == 60:
        h += 1; m = 0
    return f"{h:02d}:{m:02d}"


def build_html_grouped(result, out_path, meta=None):
    gen = datetime.now().strftime("%Y-%m-%d %H:%M")
    src_map = {"haversine": "直線估算", "google": "Google Maps 真實道路",
               "osrm": "OSRM 真實道路", "fallback": "直線估算(降級)"}
    src = src_map.get(result.distance_source, result.distance_source)
    meta = meta or {}
    start = meta.get("start_hour", 8.0)

    cards = ""
    for rt in result.routes:
        v = rt["vehicle"]
        ret = _hhmm(rt["end_hour"])
        on_time = rt.get("on_time", True)
        ret_tag = '<span class="ok">✅ 準時回倉</span>' if on_time else f'<span class="warn">⚠ 超過 17:30（{ret}）</span>'
        fuel_txt = f" ｜ 油資 {rt.get('fuel_cost', 0):.0f} 元" if result.fuel_cost_per_km > 0 else ""
        rows_html = ""
        for si, s in enumerate(rt["stops"]):
            a, lv = rt["etas"][si]
            qty = int(s.demand) if s.demand == int(s.demand) else s.demand
            item_txt = _items_str(s)
            rows_html += (
                f"<tr><td>{si+1}</td><td><b>{s.name}</b><br><span class='addr'>{s.address}</span></td>"
                f"<td class='num'>{qty}</td>"
                f"<td class='num'>{item_txt or '—'}</td>"
                f"<td class='num'>{_hhmm(a)}</td><td class='num'>{_hhmm(lv)}</td></tr>"
            )
        cards += f"""
        <div class="card">
          <div class="card-h">{v.id} 路線</div>
          <div class="card-meta">起點 {v.start_addr or '—'} ｜ {len(rt['stops'])} 站 ｜
            實際里程 {rt['distance_km']:.1f} km ｜ 總瓶數 {rt['load']:.0f} ｜
            預計回到起點 {ret}（目標 17:30）{ret_tag}{fuel_txt}</div>
          <table>
            <thead><tr><th>#</th><th>店家 / 地址</th><th>瓶數</th><th>品項</th><th>到店</th><th>離店</th></tr></thead>
            <tbody>{rows_html}</tbody>
          </table>
        </div>"""

    skipped = getattr(result, "skipped", [])
    skip_html = ""
    if skipped:
        items = "".join(f"<li>{name} — {reason}</li>" for name, reason in skipped)
        skip_html = f"<div class='skip'><b>⚠ 跳過 {len(skipped)} 筆（地址無法定位）：</b><ul>{items}</ul></div>"

    html = f"""<!DOCTYPE html>
<html lang="zh-TW"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>鮮奶配送路線報表</title>
<style>
 body{{font-family:-apple-system,"Microsoft JhengHei",sans-serif;margin:0;background:#f4f5f7;color:#222;}}
 .wrap{{max-width:960px;margin:0 auto;padding:18px;}}
 header{{background:#0b6b3a;color:#fff;padding:16px 18px;border-radius:10px;}}
 header h1{{margin:0;font-size:20px;}}
 header .sub{{opacity:.9;font-size:13px;margin-top:4px;}}
 .summary{{display:flex;gap:10px;flex-wrap:wrap;margin:14px 0;}}
 .kpi{{background:#fff;border-radius:10px;padding:10px 14px;flex:1;min-width:120px;box-shadow:0 1px 3px rgba(0,0,0,.08);}}
 .kpi .v{{font-size:20px;font-weight:700;}}
 .kpi .l{{font-size:12px;color:#666;}}
 .card{{background:#fff;border-radius:10px;padding:12px 14px;margin-bottom:12px;box-shadow:0 1px 3px rgba(0,0,0,.08);}}
 .card-h{{font-weight:700;font-size:15px;margin-bottom:4px;}}
 .card-meta{{font-size:12px;color:#555;margin-bottom:8px;}}
 table{{width:100%;border-collapse:collapse;font-size:13px;}}
 th,td{{text-align:left;padding:6px 6px;border-bottom:1px solid #eee;}}
 th{{color:#666;font-weight:600;}}
 td.num,th.num{{text-align:right;font-variant-numeric:tabular-nums;}}
 .addr{{color:#888;font-size:11px;}}
 .skip{{background:#fff3cd;border:1px solid #ffe69c;border-radius:10px;padding:10px 14px;font-size:13px;}}
 .skip ul{{margin:6px 0 0;padding-left:18px;}}
 footer{{text-align:center;color:#999;font-size:12px;margin:16px 0;}}
</style></head>
<body><div class="wrap">
<header><h1>🥛 鮮奶配送路線報表</h1>
<div class="sub">出發 {_hhmm(start)} ｜ 距離來源 {src} ｜ 生成 {gen}</div></header>
{_PNG_BTN}
<div class="summary">
  <div class="kpi"><div class="v">{len(result.routes)}</div><div class="l">出車數</div></div>
  <div class="kpi"><div class="v">{sum(len(r['stops']) for r in result.routes)}</div><div class="l">配送店數</div></div>
  <div class="kpi"><div class="v">{result.total_distance_km:.0f}</div><div class="l">總實際里程 km</div></div>
  <div class="kpi"><div class="v">{result.total_load:.0f}</div><div class="l">總瓶數</div></div>
  {f'<div class="kpi"><div class="v">{result.total_fuel_cost:.0f}</div><div class="l">預估總油資 元</div></div>' if result.fuel_cost_per_km > 0 else ''}
</div>
{cards}
{skip_html}
<footer>本報表由物流路線規劃 Agent 產出 · 下貨時間按每瓶 15 秒計算</footer>
</div>{_PNG_SCRIPT}</body></html>"""
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(html)
    return out_path


def _safe_veh(vid):
    return "".join(c if (c.isalnum() or c in "-_") else "_" for c in str(vid))


def build_html_per_vehicle(result, day_dir, meta=None):
    """每台車一個獨立 HTML(含 PNG 下載按鈕)，給不同司機分別拿。
    回傳 [(vehicle_id, filepath), ...]。"""
    meta = meta or {}
    start = meta.get("start_hour", 8.0)
    gen = datetime.now().strftime("%Y-%m-%d %H:%M")
    out = []
    css = ("<style>body{font-family:-apple-system,'Microsoft JhengHei',sans-serif;margin:0;"
           "background:#f4f5f7;color:#222;}.wrap{max-width:720px;margin:0 auto;padding:18px;}"
           "header{background:#0b6b3a;color:#fff;padding:16px 18px;border-radius:10px;}"
           "header h1{margin:0;font-size:20px;}header .sub{opacity:.9;font-size:13px;margin-top:4px;}"
           ".card{background:#fff;border-radius:10px;padding:12px 14px;margin:12px 0;"
           "box-shadow:0 1px 3px rgba(0,0,0,.08);}.card-meta{font-size:12px;color:#555;margin-bottom:8px;}"
           "table{width:100%;border-collapse:collapse;font-size:14px;}"
           "th,td{text-align:left;padding:7px 6px;border-bottom:1px solid #eee;}"
           "th{color:#666;font-weight:600;}td.num,th.num{text-align:right;font-variant-numeric:tabular-nums;}"
           ".addr{color:#888;font-size:11px;}.ok{color:#0b8a43;font-weight:700;}"
           ".warn{color:#c62828;font-weight:700;}footer{text-align:center;color:#999;font-size:12px;margin:16px 0;}"
           "@media print{.dlbtn{display:none;}}</style>")
    for rt in result.routes:
        v = rt["vehicle"]
        ret = _hhmm(rt["end_hour"])
        on_time = rt.get("on_time", True)
        ret_tag = ('<span class="ok">\u2705 \u6e96\u6642\u56de\u5009</span>' if on_time
                   else '<span class="warn">\u26a0 \u8d85\u904e 17:30 (' + ret + ')</span>')
        fuel_txt = ""
        if result.fuel_cost_per_km > 0:
            fuel_txt = " \uff5c \u6cb9\u8cc7 " + ("%.0f" % rt.get("fuel_cost", 0)) + " \u5143"
        rows = ""
        for si, sp in enumerate(rt["stops"]):
            a, lv = rt["etas"][si]
            qty = int(sp.demand) if sp.demand == int(sp.demand) else sp.demand
            rows += ("<tr><td>" + str(si + 1) + "</td><td><b>" + str(sp.name) +
                     "</b><br><span class='addr'>" + str(sp.address) + "</span></td>"
                     "<td class='num'>" + (_items_str(sp) or "—") + "</td>"
                     "<td class='num'>" + str(qty) + "</td><td class='num'>" + _hhmm(a) +
                     "</td><td class='num'>" + _hhmm(lv) + "</td></tr>")
        html = ("<!DOCTYPE html><html lang='zh-TW'><head><meta charset='utf-8'>"
                "<meta name='viewport' content='width=device-width, initial-scale=1'>"
                "<title>" + str(v.id) + " \u914d\u9001\u5831\u8868</title>" + css + "</head>"
                "<body><div class='wrap'>"
                "<header><h1>\U0001f69a " + str(v.id) + " \u914d\u9001\u5831\u8868</h1>"
                "<div class='sub'>\u51fa\u767c " + _hhmm(start) + " \uff5c " + str(len(rt["stops"])) +
                " \u7ad9 \uff5c \u91cc\u7a0b " + ("%.1f" % rt["distance_km"]) + " km \uff5c \u751f\u6210 " + gen +
                "</div></header>" + _PNG_BTN +
                "<div class='card'><div class='card-meta'>\u8d77\u9ede " + (v.start_addr or "\u2014") +
                " \uff5c \u7e3d\u74f6\u6578 " + ("%.0f" % rt["load"]) +
                " \uff5c \u9810\u8a08\u56de\u5230\u8d77\u9ede " + ret + "\uff08\u76ee\u6a19 17:30\uff09" +
                ret_tag + fuel_txt + "</div><table><thead><tr><th>#</th><th>\u5e97\u5bb6 / \u5730\u5740</th>"
                "<th class='num'>瓶數</th><th class='num'>品項</th><th class='num'>到店</th><th class='num'>離店</th></tr></thead>"
                "<tbody>" + rows + "</tbody></table></div>"
                "<footer>\u672c\u8eca\u5831\u8868\u7531\u7269\u6d41\u8def\u7dda\u898f\u5283 Agent \u7522\u51fa</footer>"
                "</div>" + _PNG_SCRIPT + "</body></html>")
        fp = os.path.join(day_dir, "route_report_" + _safe_veh(v.id) + ".html")
        with open(fp, "w", encoding="utf-8") as f:
            f.write(html)
        out.append((v.id, fp))
    return out


def build_csv_grouped(result, out_path):
    with open(out_path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(["車號", "序號", "店家", "地址", "瓶數", "品項", "下貨秒數", "預計到店", "預計離店"])
        for rt in result.routes:
            v = rt["vehicle"]
            for si, s in enumerate(rt["stops"]):
                a, lv = rt["etas"][si]
                qty = int(s.demand) if s.demand == int(s.demand) else s.demand
                svc = int(round(s.service_time))
                w.writerow([v.id, si + 1, s.name, s.address, qty, _items_str(s), svc, _hhmm(a), _hhmm(lv)])
    # 總計補一張摘要表(同檔下方另寫會蓋掉，這裡用第二種方式：寫入 summary 區)
    with open(out_path, "a", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow([])
        w.writerow(["=== 路線總計 ==="])
        w.writerow(["出車數", len(result.routes)])
        w.writerow(["總實際里程_km", f"{result.total_distance_km:.1f}"])
        w.writerow(["總瓶數", f"{result.total_load:.0f}"])
        if result.fuel_cost_per_km > 0:
            w.writerow(["油資單價_元每km", f"{result.fuel_cost_per_km:.1f}"])
            w.writerow(["預估總油資_元", f"{result.total_fuel_cost:.0f}"])
        for rt in result.routes:
            ret = _hhmm(rt["end_hour"])
            status = "準時回倉" if rt.get("on_time", True) else f"超過17:30({ret})"
            w.writerow([rt["vehicle"].id, "實際里程_km", f"{rt['distance_km']:.1f}",
                        "回倉", ret, status,
                        f"油資_{rt.get('fuel_cost',0):.0f}元" if result.fuel_cost_per_km > 0 else ""])
    return out_path


# ---------- 派車單（每台車一份，給司機/內勤） ----------

def build_dispatch_grouped(result, day_dir, meta=None):
    """產『當日派車單』：每台車一份派遣清單（HTML + 合併 CSV），放到 day_dir。
    與路線規劃總報表(當日車輛報表)分開，專給司機拿著跑。"""
    meta = meta or {}
    start = meta.get("start_hour", 8.0)
    gen = datetime.now().strftime("%Y-%m-%d %H:%M")

    # CSV：所有車攤平
    csvp = os.path.join(day_dir, "dispatch.csv")
    with open(csvp, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(["車號", "順序", "店家", "地址", "瓶數", "品項", "下貨秒數", "預計到店", "預計離店"])
        for rt in result.routes:
            v = rt["vehicle"]
            for si, s in enumerate(rt["stops"]):
                a, lv = rt["etas"][si]
                qty = int(s.demand) if s.demand == int(s.demand) else s.demand
                svc = int(round(s.service_time))
                w.writerow([v.id, si + 1, s.name, s.address, qty, _items_str(s), svc, _hhmm(a), _hhmm(lv)])
        w.writerow([])
        w.writerow(["=== 派車單總計 ==="])
        w.writerow(["出車數", len(result.routes)])
        w.writerow(["總瓶數", f"{result.total_load:.0f}"])
        if getattr(result, "fuel_cost_per_km", 0) > 0:
            w.writerow(["油資單價_元每km", f"{result.fuel_cost_per_km:.1f}"])
            w.writerow(["預估總油資_元", f"{result.total_fuel_cost:.0f}"])
        for rt in result.routes:
            ret = _hhmm(rt["end_hour"])
            status = "準時回倉" if rt.get("on_time", True) else f"超過17:30({ret})"
            fuel = f"油資_{rt.get('fuel_cost',0):.0f}元" if getattr(result, "fuel_cost_per_km", 0) > 0 else ""
            w.writerow([rt["vehicle"].id, "里程_km", f"{rt['distance_km']:.1f}", "回倉", ret, status, fuel])

    # HTML：每台車一張卡片，司機視角
    cards = ""
    for rt in result.routes:
        v = rt["vehicle"]
        ret = _hhmm(rt["end_hour"])
        on_time = rt.get("on_time", True)
        tag = '<span class="ok">✅ 準時回倉</span>' if on_time else f'<span class="warn">⚠ 超過17:30（{ret}）</span>'
        fuel_txt = f" ｜ 油資 {rt.get('fuel_cost', 0):.0f} 元" if getattr(result, "fuel_cost_per_km", 0) > 0 else ""
        rows_html = ""
        for si, s in enumerate(rt["stops"]):
            a, lv = rt["etas"][si]
            qty = int(s.demand) if s.demand == int(s.demand) else s.demand
            item_txt = _items_str(s)
            rows_html += (
                f"<tr><td class='seq'>{si+1}</td><td><b>{s.name}</b><br><span class='addr'>{s.address}</span></td>"
                f"<td class='num'>{qty}</td>"
                f"<td class='num'>{item_txt or '—'}</td>"
                f"<td class='num'>{_hhmm(a)}</td><td class='num'>{_hhmm(lv)}</td></tr>"
            )
        cards += f"""
        <div class="card">
          <div class="card-h">🚚 {v.id} 派車單</div>
          <div class="card-meta">起點 {v.start_addr or '—'} ｜ {len(rt['stops'])} 站 ｜
            總瓶數 {rt['load']:.0f} ｜ 里程 {rt['distance_km']:.1f} km ｜ 回倉 {ret}（目標17:30）{tag}{fuel_txt}</div>
          <table>
            <thead><tr><th>#</th><th>店家 / 地址</th><th>瓶數</th><th>品項</th><th>到店</th><th>離店</th></tr></thead>
            <tbody>{rows_html}</tbody>
          </table>
        </div>"""

    html = f"""<!DOCTYPE html>
<html lang="zh-TW"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>鮮奶配送派車單</title>
<style>
 body{{font-family:-apple-system,"Microsoft JhengHei",sans-serif;margin:0;background:#f4f5f7;color:#222;}}
 .wrap{{max-width:960px;margin:0 auto;padding:18px;}}
 header{{background:#c0392b;color:#fff;padding:16px 18px;border-radius:10px;}}
 header h1{{margin:0;font-size:20px;}}
 header .sub{{opacity:.9;font-size:13px;margin-top:4px;}}
 .card{{background:#fff;border-radius:10px;padding:12px 14px;margin-bottom:12px;box-shadow:0 1px 3px rgba(0,0,0,.08);}}
 .card-h{{font-weight:700;font-size:16px;margin-bottom:4px;}}
 .card-meta{{font-size:12px;color:#555;margin-bottom:8px;}}
 table{{width:100%;border-collapse:collapse;font-size:13px;}}
 th,td{{text-align:left;padding:6px 6px;border-bottom:1px solid #eee;}}
 th{{color:#666;font-weight:600;}}
 td.num,th.num{{text-align:right;font-variant-numeric:tabular-nums;}}
 td.seq{{font-weight:700;color:#c0392b;font-size:15px;width:28px;}}
 .addr{{color:#888;font-size:11px;}}
 .ok{{color:#0b6b3a;font-weight:700;}}
 .warn{{color:#c0392b;font-weight:700;}}
 footer{{text-align:center;color:#999;font-size:12px;margin:16px 0;}}
</style></head>
<body><div class="wrap">
<header><h1>🚚 鮮奶配送派車單</h1>
<div class="sub">出發 {_hhmm(start)} ｜ 生成 {gen} ｜ 下貨每瓶 15 秒{f" ｜ 預估總油資 {result.total_fuel_cost:.0f} 元（{result.fuel_cost_per_km:.1f} 元/km）" if getattr(result, "fuel_cost_per_km", 0) > 0 else ""}</div></header>
{cards}
<footer>本派車單由物流路線規劃 Agent 產出 · 每台車一份，司機拿著跑</footer>
</div></body></html>"""
    htmlp = os.path.join(day_dir, "dispatch.html")
    with open(htmlp, "w", encoding="utf-8") as f:
        f.write(html)
    return htmlp, csvp


# ---------- 結構化資料（供客服助理撈取 ETA/貨品） ----------

def build_dispatch_data(result, out_path, meta=None):
    """產 dispatch_data.json：每店貨品/數量/ETA/所屬車 + 每台車載貨量。
    供客服助理讀取後通知客戶。回傳 out_path。"""
    meta = meta or {}
    start = meta.get("start_hour", 9.5)
    data = {
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "start_hour": start,
        "vehicles": [],
        "stops": [],
    }
    for rt in result.routes:
        v = rt["vehicle"]
        # 載貨量彙總
        agg = {}
        milk_total = 0.0
        for s in rt["stops"]:
            milk_total += (s.demand or 0)
            for n, d in (getattr(s, "items", None) or {}).items():
                if n in agg:
                    agg[n]["qty"] += d["qty"]
                else:
                    agg[n] = dict(d)
        manifest = [{"item": "鮮奶", "qty": round(milk_total), "unit": "瓶"}]
        manifest += [{"item": n, "qty": round(d["qty"]), "unit": d["unit"]}
                     for n, d in agg.items()]
        data["vehicles"].append({
            "vehicle": v.id,
            "start_addr": v.start_addr or "",
            "stops_count": len(rt["stops"]),
            "distance_km": round(rt["distance_km"], 1),
            "load_bottles": round(rt["load"], 0),
            "return_hour": _hhmm(rt["end_hour"]),
            "on_time": rt.get("on_time", True),
            "manifest": manifest,
        })
        for si, s in enumerate(rt["stops"]):
            a, lv = rt["etas"][si]
            items = [{"item": n, "qty": round(d["qty"]), "unit": d["unit"]}
                     for n, d in (getattr(s, "items", None) or {}).items()]
            data["stops"].append({
                "vehicle": v.id,
                "seq": si + 1,
                "name": s.name,
                "address": getattr(s, "address", ""),
                "bottles": round(s.demand) if s.demand == int(s.demand) else s.demand,
                "items": items,
                "service_sec": int(round(getattr(s, "service_time", 0) or 0)),
                "arrive": _hhmm(a),
                "leave": _hhmm(lv),
            })
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    return out_path


# ---------- 整合 Excel（一份 xlsx，3 分頁） ----------

def build_workbook(result, out_path, meta=None, map_png=None):
    """產整合 Excel：①路線總表 ②各車派車單 ③油資/里程總計 ④路線圖(可選 PNG)。
    map_png: 路線圖 PNG 路徑；None 則不加路線圖分頁。回傳 out_path。"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.utils import get_column_letter

    meta = meta or {}
    start = meta.get("start_hour", 9.5)
    has_fuel = getattr(result, "fuel_cost_per_km", 0) > 0
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="0B6B3A")
    veh_font = Font(bold=True, color="FFFFFF")
    veh_fill = PatternFill("solid", fgColor="C0392B")
    center = Alignment(horizontal="center")

    def _style_header(ws, row=1):
        for c in ws[row]:
            c.font = hdr_font; c.fill = hdr_fill; c.alignment = center

    def _autofit(ws):
        for col in ws.columns:
            w = max((len(str(c.value)) for c in col if c.value is not None), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max(w + 2, 8), 48)

    wb = openpyxl.Workbook()

    # ① 路線總表
    # ① 路線總表
    ws1 = wb.active
    h1 = ["車號", "序號", "店家", "地址", "瓶數", "品項", "下貨秒數", "預計到店", "預計離店"]
    ws1.append(h1)
    for rt in result.routes:
        v = rt["vehicle"]
        for si, s in enumerate(rt["stops"]):
            a, lv = rt["etas"][si]
            qty = int(s.demand) if s.demand == int(s.demand) else s.demand
            ws1.append([v.id, si + 1, s.name, s.address, qty, _items_str(s),
                        int(round(s.service_time)), _hhmm(a), _hhmm(lv)])
    _style_header(ws1); ws1.freeze_panes = "A2"; _autofit(ws1)

    # ② 各車派車單（每台車一段，段首車號列）
    ws2 = wb.create_sheet("各車派車單")
    h2 = ["車號/資訊", "序號", "店家", "地址", "瓶數", "到店", "離店"]
    ws2.append(h2)
    for rt in result.routes:
        v = rt["vehicle"]
        ret = _hhmm(rt["end_hour"])
        status = "準時回倉" if rt.get("on_time", True) else f"超過17:30({ret})"
        fuel = f"｜油資 {rt.get('fuel_cost', 0):.0f} 元" if has_fuel else ""
        info = (f"🚚 {v.id}｜起點 {v.start_addr or '—'}｜{len(rt['stops'])}站"
                f"｜{rt['load']:.0f}瓶｜{rt['distance_km']:.1f}km｜回倉 {ret}（{status}）{fuel}")
        r = ws2.max_row + 1
        ws2.append([info])
        for c in ws2[r]:
            c.font = veh_font
        ws2[r][0].fill = veh_fill
        ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        for si, s in enumerate(rt["stops"]):
            a, lv = rt["etas"][si]
            qty = int(s.demand) if s.demand == int(s.demand) else s.demand
            ws2.append(["", si + 1, s.name, s.address, qty, _hhmm(a), _hhmm(lv)])
    _style_header(ws2); ws2.freeze_panes = "A2"; _autofit(ws2)

    # ③ 油資/里程總計
    ws3 = wb.create_sheet("總計")
    ws3.append(["項目", "數值"])
    ws3.append(["出車數", len(result.routes)])
    ws3.append(["配送店數", sum(len(r["stops"]) for r in result.routes)])
    ws3.append(["總瓶數", f"{result.total_load:.0f}"])
    ws3.append(["總實際里程_km", f"{result.total_distance_km:.1f}"])
    if has_fuel:
        ws3.append(["油資單價_元每km", f"{result.fuel_cost_per_km:.1f}"])
        ws3.append(["預估總油資_元", f"{result.total_fuel_cost:.0f}"])
    ws3.append([])
    ws3.append(["車號", "站數", "里程_km", "瓶數", "回倉", "狀態"] + (["油資_元"] if has_fuel else []))
    for rt in result.routes:
        ret = _hhmm(rt["end_hour"])
        status = "準時回倉" if rt.get("on_time", True) else f"超過17:30({ret})"
        row = [rt["vehicle"].id, len(rt["stops"]), f"{rt['distance_km']:.1f}",
               f"{rt['load']:.0f}", ret, status]
        if has_fuel:
            row.append(f"{rt.get('fuel_cost', 0):.0f}")
        ws3.append(row)
    _style_header(ws3); _autofit(ws3)

    # ③b 每台車載貨量（manifest，司機早上撿貨用）
    ws3b = wb.create_sheet("每台車載貨量")
    ws3b.append(["車號", "品項", "數量", "單位"])
    for rt in result.routes:
        v = rt["vehicle"]
        # 彙總該車所有站的非鮮奶品項 + 鮮奶總瓶數
        agg = {}  # 品名 -> {"qty":float,"unit":str}
        milk_total = 0.0
        for s in rt["stops"]:
            milk_total += (s.demand or 0)
            for n, d in (getattr(s, "items", None) or {}).items():
                if n in agg:
                    agg[n]["qty"] += d["qty"]
                else:
                    agg[n] = dict(d)
        # 車號小計標頭
        r0 = ws3b.max_row + 1
        ws3b.append([f"🚚 {v.id}｜共 {len(rt['stops'])} 站｜鮮奶 {milk_total:.0f} 瓶", "", "", ""])
        for c in ws3b[r0]:
            c.font = veh_font
        ws3b[r0][0].fill = veh_fill
        ws3b.append(["", "鮮奶(瓶數)", f"{milk_total:.0f}", "瓶"])
        for n, d in agg.items():
            ws3b.append(["", n, f"{d['qty']:.0f}", d["unit"]])
        ws3b.append([])
    _style_header(ws3b); _autofit(ws3b)

    # ④ 路線圖（若有 PNG）
    if map_png and os.path.exists(map_png):
        ws4 = wb.create_sheet("路線圖")
        try:
            img = XLImage(map_png)
            # 限制寬度，避免過大；保持比例
            max_w = 1100
            if img.width > max_w:
                ratio = max_w / img.width
                img.width = max_w
                img.height = int(img.height * ratio)
            ws4.add_image(img, "A1")
            ws4.sheet_view.showGridLines = False
        except Exception as e:
            ws4.append([f"路線圖載入失敗：{e}"])

    wb.save(out_path)
    return out_path
