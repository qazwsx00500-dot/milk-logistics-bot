"""
data_loader.py — 讀取店家資料 (Excel / CSV)，照車號分組

預期欄位（中文或英文皆可）：
  車號       / 車輛 / 路線編號 / route / vehicle      -> 分組依據
  店家名稱   / 名稱 / 店名 / name
  店家地址   / 地址 / address
  瓶數       / 數量 / qty / bottles
  品項       / 品名 / 貨品 / 項目 / item
                -> 非鮮奶貨品，文字格式「品名數量(單位)」可用逗號隔多品項
                   (例如: 冰勃朗非氫化基底乳1(箱),鳳梨果泥3(包))
                   鮮乳類已計入「瓶數」，不在此欄
  出發點地址 / 起點地址 / 倉庫地址 / start_address / depot_address  -> 每台車出發點

自動：
  - 地理編碼 (店家地址 + 出發點地址 -> 座標)
  - 由瓶數計算 下貨時間 (瓶數 * 15 秒)
  - 若有非鮮奶品項，每店額外加 180 秒 (約3分鐘)
  - 照車號分組，每台車獨立回傳

回傳 (vehicles, stops_by_vehicle, skipped)
  vehicles: [Vehicle]           每台的起點座標
  stops_by_vehicle: {車號: [Stop]}
  skipped: [(名稱, 原因)]
"""

import os
import csv
import re

try:
    import openpyxl
except ImportError:
    openpyxl = None

from geocoder import geocode
from route_planner import Stop, Vehicle

SERVICE_SEC_PER_BOTTLE = 15.0


def _norm_header(h):
    return (h or "").strip().lower().replace(" ", "").replace("　", "")


_VEH_KEYS = ["車號", "車輛", "路線編號", "路線", "route", "vehicle", "車"]
_NAME_KEYS = ["店家名稱", "名稱", "店名", "客戶名稱", "name", "店家"]
_ADDR_KEYS = ["店家地址", "地址", "客戶地址", "address", "addr"]
_QTY_KEYS = ["瓶數", "數量", "箱數", "瓶量", "qty", "bottles", "count"]
_ITEM_KEYS = ["品項", "品名", "貨品", "品項名稱", "項目", "item", "items"]
_CONS_KEYS = ["特殊需求", "特殊要求", "需求", "備註", "備註說明", "constraint", "note", "remark"]
EXTRA_SERVICE_SEC_FOR_ITEMS = 180.0   # 該店有非鮮奶品項時，額外加 180 秒 (~3分)

# ── 品名 → 強制單位對照表（Ann 確認規則，2026-07-19）──────────────
#   保久乳 / 糖漿      → 強制「瓶」（無視 Excel 原單位）
#   紫米紅豆           → 強制「罐」
#   鳳梨果泥           → 強制「包」
#   冰勃朗             → 不列入（維持 Excel 原單位，瓶或箱依表為主）
#   其餘未列品項       → 不列入（維持 Excel 原單位，空則回退「件」）
# 匹配方式：品名「包含」關鍵字即命中（如「冰勃朗非氫化基底乳」含「冰勃朗」）。
UNIT_OVERRIDE = {
    "保久乳": "瓶",
    "糖漿": "瓶",
    "紫米紅豆": "罐",
    "鳳梨果泥": "包",
}

def _override_unit(name, unit):
    """品名命中對照表則強制覆寫單位，否則維持原 unit（空則回退『件』）。"""
    for kw, u in UNIT_OVERRIDE.items():
        if kw in name:
            return u
    return unit
_FRESH_MILK_HINTS = ["鮮乳", "鮮奶", "牛奶"]
_START_KEYS = ["出發點地址", "起點地址", "倉庫地址", "出發地址",
               "start_address", "depot_address", "start", "depot"]
_FUEL_KEYS = ["油資單價", "油資", "油錢單價", "fuel", "fuel_cost", "fuel_cost_per_km", "元每km"]


def _read_fuel_from_rows(rows, headers):
    """從資料列讀『油資單價』欄：回傳第一個有效值(float)，全空則回 None。全車共用。"""
    idx_by_norm = {_norm_header(h): h for h in headers}
    for row in rows:
        v = _get(row, _FUEL_KEYS, idx_by_norm)
        if v not in (None, ""):
            try:
                f = float(v)
                if f > 0:
                    return f
            except (ValueError, TypeError):
                pass
    return None


def read_fuel_cost(path):
    """從資料檔(xlsx/csv)讀『油資單價』欄，回傳 float 或 None（供 main 覆蓋 .env）。"""
    try:
        if str(path).lower().endswith(".csv"):
            with open(path, encoding="utf-8-sig", newline="") as f:
                reader = csv.DictReader(f)
                return _read_fuel_from_rows([dict(r) for r in reader], reader.fieldnames or [])
        if openpyxl is None:
            return None
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        rows_raw = list(ws.iter_rows(values_only=True))
        if not rows_raw:
            return None
        headers = [str(h) if h is not None else "" for h in rows_raw[0]]
        rows = [{headers[i]: r[i] for i in range(len(headers))}
                for r in rows_raw[1:] if not all(v is None for v in r)]
        return _read_fuel_from_rows(rows, headers)
    except Exception:
        return None


def _get(row, keys, idx_by_norm):
    for k in keys:
        orig = idx_by_norm.get(k)
        if orig is not None and orig in row:
            return row[orig]
    return None


def parse_constraint(raw):
    """把『特殊需求』欄文字解析成約束 dict。
    支援語法（多條用 中文/英文 逗號、頓號、分號分隔）：
      時間窗上界(必須幾點『前』到):  '10前' '上午10前' '10:30前' '10點前' '10:00前' 'AM10前'
      時間窗下界(必須幾點『後』到):  '14後' '下午2後' '14:00後' '2點後' 'PM2後'
      區間(只能這段送):              '10-14' '10:00-14:00' '10~14'
      固定首站:                      '首站' '第一站' '優先' '先送'
      固定末站:                      '末站' '最後站' '最後' '後送' '晚送'
    回傳 {"time_lb":float|None, "time_ub":float|None, "first":bool, "last":bool, "raw":str}
    無效/空白 -> {}（表示無約束）。
    """
    out = {"time_lb": None, "time_ub": None, "first": False, "last": False, "raw": ""}
    if not raw or not str(raw).strip():
        return out
    txt = str(raw).strip()
    out["raw"] = txt
    # 先抓區間 10-14 / 10:00-14:00 / 10~14
    m_rng = re.search(r"(\d{1,2})(?::(\d{2}))?\s*[-\uFF5E~]\s*(\d{1,2})(?::(\d{2}))?", txt)
    if m_rng:
        lb_h, lb_m, ub_h, ub_m = m_rng.groups()
        out["time_lb"] = int(lb_h) + (int(lb_m) if lb_m else 0) / 60.0
        out["time_ub"] = int(ub_h) + (int(ub_m) if ub_m else 0) / 60.0
        # 區間已涵蓋，不再個別抓前/後（但首末站仍可共存）
    # 抓「前」(上界)
    m_ub = re.search(r"(?:上午|下午|早上|am|pm)?\s*(\d{1,2})(?::(\d{2}))?\s*點?\s*前", txt, re.IGNORECASE)
    if m_ub and out["time_ub"] is None:
        h, m = m_ub.groups()
        hh = int(h)
        if re.search("下午|pm", txt, re.IGNORECASE) and hh < 12:
            hh += 12
        out["time_ub"] = hh + (int(m) if m else 0) / 60.0
    # 抓「後」(下界)
    m_lb = re.search(r"(?:上午|下午|早上|am|pm)?\s*(\d{1,2})(?::(\d{2}))?\s*點?\s*後", txt, re.IGNORECASE)
    if m_lb and out["time_lb"] is None:
        h, m = m_lb.groups()
        hh = int(h)
        if re.search("下午|pm", txt, re.IGNORECASE) and hh < 12:
            hh += 12
        out["time_lb"] = hh + (int(m) if m else 0) / 60.0
    # 首站
    if re.search("首站|第一站|優先|先送|最早", txt):
        out["first"] = True
    # 末站
    if re.search("末站|最後站|最後|後送|晚送", txt):
        out["last"] = True
    return out


def parse_items(raw):
    """把『品項』欄文字解析成 {品名: {"qty": float, "unit": str}}。
    格式: 品名數量(單位)，多品項用逗號/、/；分隔。例:
      '冰勃朗非氫化基底乳1(箱),鳳梨果泥3(包)'
    數量可能帶小數；單位可省略(預設 '件')。空白或 None -> {}。"""
    out = {}
    if not raw or not str(raw).strip():
        return out
    for part in re.split(r"[,，、;；]+", str(raw)):
        part = part.strip()
        if not part:
            continue
        m = re.match(r"^(.+?)\s*[-]?\s*([\d.]+)\s*(?:\(([^)]*)\))?\s*$", part)
        if m:
            name = m.group(1).strip()
            try:
                qty = float(m.group(2))
            except ValueError:
                qty = 0.0
            unit = (m.group(3) or "").strip() or "件"
            unit = _override_unit(name, unit)
        else:
            name = part
            qty = 1.0
            unit = "件"
            unit = _override_unit(name, unit)
        if name:
            if name in out:
                out[name]["qty"] += qty
            else:
                out[name] = {"qty": qty, "unit": unit}
    return out


def load_from_rows(rows, headers, depot=None):
    idx_by_norm = {_norm_header(h): h for h in headers}
    vehicles = {}            # 車號 -> Vehicle (暫存)
    stops_by_vehicle = {}    # 車號 -> [Stop]
    skipped = []
    geo_jobs = []            # [(label, address, callback)]

    for i, row in enumerate(rows):
        veh_raw = _get(row, _VEH_KEYS, idx_by_norm)
        name = str(_get(row, _NAME_KEYS, idx_by_norm) or f"店家{i+1}").strip()
        addr = str(_get(row, _ADDR_KEYS, idx_by_norm) or "").strip()
        start_raw = _get(row, _START_KEYS, idx_by_norm)
        start_addr = str(start_raw or "").strip() if start_raw is not None else ""
        qty_raw = _get(row, _QTY_KEYS, idx_by_norm)
        try:
            qty = float(qty_raw) if qty_raw not in (None, "") else 0.0
        except ValueError:
            qty = 0.0
        # 品項欄：非鮮奶貨品，文字 -> dict
        item_raw = _get(row, _ITEM_KEYS, idx_by_norm)
        cons_raw = _get(row, _CONS_KEYS, idx_by_norm)
        cons = parse_constraint(cons_raw) if cons_raw else {}
        items = parse_items(item_raw)
        svc = qty * SERVICE_SEC_PER_BOTTLE
        if items:
            svc += EXTRA_SERVICE_SEC_FOR_ITEMS   # 有非鮮奶品項 → 每店 +3分

        veh = str(veh_raw).strip() if veh_raw not in (None, "") else "未分車"
        if not addr:
            skipped.append((name, "缺少店家地址"))
            continue

        stop = Stop(id=f"{veh}-{len(stops_by_vehicle.get(veh, []))+1:03d}",
                    name=name, lat=0.0, lon=0.0, demand=qty,
                    service_time=svc, address=addr, vehicle=veh,
                    items=dict(items))
        if cons:
            stop.constraint = cons
        stops_by_vehicle.setdefault(veh, []).append(stop)
        # 起點：若有總倉 depot 則統一用總倉；否則用 Excel 出發點地址欄
        if depot is not None:
            vehicles.setdefault(veh, {"name": veh, "depot": True,
                                      "start_addr": depot.address,
                                      "lat": depot.lat, "lon": depot.lon})
        else:
            if veh not in vehicles:
                vehicles[veh] = {"name": veh, "start_addr": start_addr}
            geo_jobs.append(("start", veh, None, start_addr))
        # 地理編碼任務（僅店家）
        geo_jobs.append(("stop", veh, len(stops_by_vehicle[veh]) - 1, addr))

    # 批量地理編碼（僅店家；depot 模式不起點）
    addrs = [job[3] for job in geo_jobs if job[0] == "stop"]
    geo_result = geocode_batch(addrs)
    addr_to_coord = dict(zip(addrs, geo_result))

    # 回填
    for kind, veh, sidx, addr in geo_jobs:
        if kind != "stop":
            continue
        coord = addr_to_coord.get(addr)
        st = stops_by_vehicle[veh][sidx]
        if coord:
            st.lat, st.lon = coord
        else:
            skipped.append((st.name, f"店家地址無法定位：{addr}"))
            st.lat = None

    # 去掉無效 stop
    for veh in list(stops_by_vehicle.keys()):
        stops_by_vehicle[veh] = [s for s in stops_by_vehicle[veh] if s.lat is not None]
        if not stops_by_vehicle[veh]:
            del stops_by_vehicle[veh]

    # 建 Vehicle 物件
    vlist = []
    for veh, info in vehicles.items():
        if info.get("lat") is None:
            sk = stops_by_vehicle.get(veh)
            if sk:
                info["lat"], info["lon"] = sk[0].lat, sk[0].lon
                info["start_addr"] = info.get("start_addr") or sk[0].address
            else:
                continue
        vlist.append(Vehicle(id=veh, name=veh,
                             start_lat=info["lat"], start_lon=info["lon"],
                             start_addr=info.get("start_addr", "")))

    return vlist, stops_by_vehicle, skipped


# 用 geocoder 的批次（含快取）
def geocode_batch(addresses):
    """一次性批次地理編碼（內部已含持久化快取：命中者零網路呼叫）。
    注意：原先寫成 [geocode_many([a]) for a in addresses] —— 對每個地址
    單獨呼叫一次，等於把批量介面包裝成「逐條串行」+ 重複快取查詢開銷，
    幾十個店家就變幾十次往返。改成一次 geocode_many(addresses) 讓快取命中者
    瞬回、未命中者統一批量處理，地理編碼階段顯著加速。"""
    from geocoder import geocode_many
    # 去重保順序（同地址只查一次，避免重複 Google 呼叫）
    uniq = list(dict.fromkeys(addresses))
    res = geocode_many(uniq)
    return [res.get(a) for a in addresses]


def load_excel(path, depot=None):
    if openpyxl is None:
        raise RuntimeError("尚未安裝 openpyxl：請執行 uv pip install openpyxl")
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows_raw = list(ws.iter_rows(values_only=True))
    if not rows_raw:
        return [], {}, []
    headers = [str(h) if h is not None else "" for h in rows_raw[0]]
    rows = []
    for r in rows_raw[1:]:
        if all(v is None for v in r):
            continue
        rows.append({headers[i]: r[i] for i in range(len(headers))})
    return load_from_rows(rows, headers, depot=depot)


def load_csv(path, depot=None):
    with open(path, encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        headers = reader.fieldnames or []
        rows = [dict(r) for r in reader]
    return load_from_rows(rows, headers, depot=depot)


def load(path, depot=None):
    ext = os.path.splitext(path)[1].lower()
    # OneDrive / Excel 開啟中常鎖住 xlsx 導致 PermissionError。
    # 先嘗試直接讀；失敗則複製到本地 temp 再讀（並重試，因鎖檔通常短暫）。
    read_path = _copy_to_temp(path)
    if ext in (".xlsx", ".xlsm"):
        return load_excel(read_path, depot=depot)
    if ext in (".csv", ".txt"):
        return load_csv(read_path, depot=depot)
    raise ValueError(f"不支援的檔案格式：{ext}")


def _copy_to_temp(path):
    """複製到本地 temp 再回傳副本路徑（繞過 OneDrive/Excel 鎖檔）。
    若原檔可直讀就直接用；否則重試複製幾次（鎖檔通常短暫），都失敗則退回原路徑。"""
    import shutil, tempfile, time
    # 先試原檔能否直開
    try:
        with open(path, "rb") as f:
            f.read(1)
        return path
    except (PermissionError, OSError):
        pass
    d = tempfile.gettempdir()
    dst = os.path.join(d, "logi_" + os.path.basename(path))
    last_err = None
    for attempt in range(5):
        try:
            shutil.copy2(path, dst)
            return dst
        except Exception as e:
            last_err = e
            time.sleep(1.5)   # 等 OneDrive/Excel 釋放鎖
    # 都失敗：嘗試原檔（讓上層報錯時能給明確訊息）
    return path
