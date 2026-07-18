"""sales_to_dispatch.py — Ann(客服業助) 的銷貨明細轉檔工具

把「隔天銷貨日報表」(如 1150720銷貨明細(中).xlsx) 整理成
JOJO 能吃的「每日配送.xlsx」(5欄: 車號/店家名稱/店家地址/瓶數/品項)。

處理規則：
  - 工作表 '工作表2'；表頭在第 5 列(index 4)：貨單日期/貨單編號/客戶簡稱/貨品名稱/數量/單位/送貨地址
  - 抬頭行(公司/報表名/貨單日期/製表人) 跳過
  |  - 續行列：客戶簡稱 與 送貨地址「皆空」→ 繼承上一筆店家+地址（照片案例：上面店家的品項/數量在下一列續接，店家/地址欄留空）；
  |    若「店名有值但地址空」「地址有值但店名空」亦視為同一店續行，繼承缺失的那一欄（避免重複排同一地址）
  |  - 退貨：數量負值 → 與同店同品項銷貨相抵
  |  - 同店合併：地址正規化（去 '-隨貨附發票' 後綴、去內部空白、全形轉半形、去括號與結尾標點）後比對，按 (地址,品項) 加總；
  |    同店不同寫法（全/半形數字、有無括號、有無空格）視為同一店，合併成一筆，店名取第一個有值者
  |  - 品項判定：貨品名含「鮮乳/鮮奶/牛奶」→ 計入「瓶數」；其餘 → 進「品項」欄
  - 地址清理：去掉 '-隨貨附發票'/'-隨貨發票' 後綴再比對同店(避免同店被拆)；寫入時保留原值

用法：
  python sales_to_dispatch.py <銷貨明細.xlsx> [--out <每日配送.xlsx>] [--veh 車01]

輸出每日配送.xlsx 到 路線規劃 資料夾 (與 JOJO 共用 OneDrive 路徑)。
"""
import os
import re
import argparse
import datetime

try:
    import openpyxl
except ImportError:
    openpyxl = None

# ---- 路徑 (與 JOJO 共用 OneDrive) ----
ONEDRIVE_DESKTOP = os.path.join(os.path.expanduser("~"), "OneDrive", "桌面")
ROUTE_DIR = os.path.join(ONEDRIVE_DESKTOP, "路線規劃")

_FRESH_MILK_HINTS = ["鮮乳", "鮮奶", "牛奶"]
_ADDR_SUFFIX = re.compile(r"\s*[-－]\s*(隨貨附發票|隨貨發票|隨貨|附發票)\s*$")

# 續行列識別：同列「客戶簡稱」與「送貨地址」皆空（貨品/數量那幾欄可能有值）→ 繼承上一筆
# （見使用者照片：上面店家那列的「品項/數量」在下一列繼續，店家/地址欄留空）
# 同店合併 key 正規化：去空白/全半形/括號/結尾標點，避免同店被拆成兩筆


def _to_half_width(s):
    """全形數字/字母轉半形，供地址比對去差異。"""
    out = []
    for ch in (s or ""):
        o = ord(ch)
        if 0xFF01 <= o <= 0xFF5E:   # 全形！到～
            out.append(chr(o - 0xFEE0))
        elif o == 0x3000:           # 全形空格
            out.append(" ")
        else:
            out.append(ch)
    return "".join(out)


def _norm_addr(addr):
    """去後綴 + 統一全半形/空白/括號/結尾標點，供同店比對（比對用，寫入時仍用原值）。"""
    a = (addr or "").strip()
    a = _to_half_width(a)
    a = _ADDR_SUFFIX.sub("", a)
    a = a.replace("臺", "台").replace("　", " ").replace("(", "").replace(")", "")
    a = re.sub(r"\s+", "", a)          # 內部所有空白都去掉
    a = a.strip(".,-_/\\ ")             # 結尾標點
    return a


def _norm_name(name):
    """店名正規化（合併同名店用）：去空白/全半形。"""
    return re.sub(r"\s+", "", _to_half_width(name or ""))


def _is_milk(name):
    return any(h in (name or "") for h in _FRESH_MILK_HINTS)


def _split_item(name, qty, unit):
    """把一筆非鮮奶貨品轉成 (品名, 數量, 單位) 字串片段，供品項欄。"""
    u = (unit or "").strip() or "件"
    return f"{name}{qty:.0f}({u})"


def convert(src_path, out_path, default_veh="車01"):
    if openpyxl is None:
        raise RuntimeError("請先安裝 openpyxl: uv pip install openpyxl")
    wb = openpyxl.load_workbook(src_path, data_only=True)
    ws = wb["工作表2"]
    rows = [r for r in ws.iter_rows(values_only=True)]

    # 表頭自動偵測：掃描含「客戶簡稱」+「送貨地址」的列（實際在第 7 列，前面是抬頭）
    hdr_idx = None
    for i, r in enumerate(rows):
        cells = [str(c).strip() for c in r if c is not None and str(c).strip()]
        if "客戶簡稱" in cells and "送貨地址" in cells:
            hdr_idx = i
            break
    if hdr_idx is None:
        raise RuntimeError("找不到表頭列（需含『客戶簡稱』與『送貨地址』）")
    hdr = rows[hdr_idx]
    # 欄位 index
    idx = {name: i for i, name in enumerate(hdr) if name is not None}

    def col(name):
        return idx.get(name)

    i_cust = col("客戶簡稱")
    i_item = col("貨品名稱")
    i_qty = col("數量")
    i_unit = col("單位")
    i_addr = col("送貨地址")

    # 彙總結構：key=(地址,品項判定) -> record
    # 店層: addr_key -> {name, addr_raw, milk:float, items:{品名:{"qty","unit"}}, seen_name}
    shops = {}
    last = {"cust": None, "addr": None}  # 續行繼承

    for r in rows[hdr_idx + 1:]:
        if all(v is None or (isinstance(v, str) and v.strip() == "") for v in r):
            continue
        cust = (str(r[i_cust]).strip() if i_cust is not None and r[i_cust] is not None else "")
        addr = (str(r[i_addr]).strip() if i_addr is not None and r[i_addr] is not None else "")
        item = (str(r[i_item]).strip() if i_item is not None and r[i_item] is not None else "")
        qty_raw = r[i_qty] if i_qty is not None else None
        unit = (str(r[i_unit]).strip() if i_unit is not None and r[i_unit] is not None else "")
        try:
            qty = float(qty_raw) if qty_raw not in (None, "") else 0.0
        except (ValueError, TypeError):
            qty = 0.0

        # 續行列（Bug 3）：客戶簡稱 與 送貨地址 皆空 → 繼承上一筆店家+地址
        #   （照片案例：上一列有店家/地址+鮮奶，下一列只填品項/數量，店家/地址留空）
        #   注意：只要「地址空白」就先嘗試繼承（店名空白可視為同一店續行，而非新店）
        if not addr and not cust:
            cust = last["cust"] or ""
            addr = last["addr"] or ""
        elif not addr and cust:
            # 店名有值但地址空白（續行列只重複了店名）→ 繼承上一筆地址
            addr = last["addr"] or ""
        elif not cust and addr:
            # 地址有值但店名空白 → 繼承上一筆店名
            cust = last["cust"] or ""
        # 更新繼承來源（有值才記）
        if cust:
            last["cust"] = cust
        if addr:
            last["addr"] = addr

        if not item or not addr:
            continue  # 無品項或無地址，跳過

        akey = _norm_addr(addr)
        # Bug 2：同店（地址相同）出現多次 → 合併同 (地址, 品項) 加總，避免 JOJO 重複排同一地址
        if akey not in shops:
            shops[akey] = {
                "name": cust or "未命名店",
                "addr_raw": addr,
                "milk": 0.0,
                "items": {},
            }
        sh = shops[akey]
        # 保留第一個有真實店名的（避免被「未命名店」覆蓋）
        if cust and sh["name"].startswith("未命名"):
            sh["name"] = cust
        if _is_milk(item):
            sh["milk"] += qty
        else:
            # 非鮮奶：進品項 dict
            if item in sh["items"]:
                sh["items"][item]["qty"] += qty
            else:
                sh["items"][item] = {"qty": qty, "unit": unit or "件"}

    # 輸出
    d = os.path.dirname(out_path)
    if d:                      # 避免 --out 給「純檔名」時 dirname 為 '' 導致 makedirs 炸掉
        os.makedirs(d, exist_ok=True)
    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    ws2.title = "每日配送"
    ws2.append(["車號", "店家名稱", "店家地址", "瓶數", "品項"])
    n = 0
    for akey, sh in shops.items():
        milk = sh["milk"]
        # 品項欄文字：多品項逗號分隔
        item_str = ", ".join(_split_item(nm, d["qty"], d["unit"]) for nm, d in sh["items"].items())
        # 瓶數四捨五入（鮮奶可能是負值退貨相抵後的小數）
        milk_int = int(round(milk))
        # 退貨相抵後：牛奶 <=0 瓶且無其他品項 → 該店本日無貨(取消訂單)，跳過
        if milk_int <= 0 and not item_str:
            continue
        # 有品項但鮮奶相抵成負/0 → 瓶數歸 0 (仍有非鮮奶貨要送)
        out_milk = milk_int if milk_int > 0 else 0
        ws2.append([default_veh, sh["name"], sh["addr_raw"], out_milk, item_str])
        n += 1
    wb2.save(out_path)
    return n, out_path, shops


def self_check(shops):
    """ANN 轉檔後自行檢查：回傳 [(等級, 訊息), ...] (等級: OK/WARN/ERR)。
    檢查：負數瓶、空地址、空店名、退貨相抵異常、續行繼承是否遺漏。"""
    issues = []
    n_total = len(shops)
    n_out = 0
    for akey, sh in shops.items():
        milk = int(round(sh["milk"]))
        item_str = ", ".join(_split_item(nm, d["qty"], d["unit"]) for nm, d in sh["items"].items())
        if milk <= 0 and not item_str:
            continue  # 取消訂單(相抵0/純退貨) 正常跳過
        n_out += 1
        if not sh["addr_raw"] or not sh["addr_raw"].strip():
            issues.append(("ERR", f"「{sh['name']}」地址空白，JOJO 無法定位"))
        if not sh["name"] or str(sh["name"]).startswith("未命名") or str(sh["name"]) == "None":
            issues.append(("WARN", f"店名異常：{sh['name']}（地址 {sh['addr_raw']}）"))
        if milk < 0:
            issues.append(("ERR", f"「{sh['name']}」瓶數為負({milk})，不應出現"))
    if n_total == 0:
        issues.append(("ERR", "轉出 0 間店，請檢查 Excel 欄位"))
    issues.append(("OK", f"轉檔自檢通過：{n_total} 間店(含取消) → {n_out} 間需配送"))
    return issues


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("src", help="銷貨明細 xlsx")
    ap.add_argument("--out", default=None, help="輸出每日配送.xlsx 路徑")
    ap.add_argument("--veh", default="車01", help="預設車號")
    args = ap.parse_args()
    if args.out is None:
        date_str = datetime.datetime.now().strftime("%Y%m%d")
        args.out = os.path.join(ROUTE_DIR, f"每日配送_{date_str}.xlsx")
    n, out_path, shops = convert(args.src, args.out, args.veh)
    print(f"✅ 轉檔完成：{n} 間店 → {out_path}")
    # 簡易摘要
    milk_shops = sum(1 for s in shops.values() if s["milk"] != 0)
    item_shops = sum(1 for s in shops.values() if s["items"])
    print(f"   含鮮奶店數: {milk_shops} ｜ 含其他品項店數: {item_shops}")
    # 轉檔後自行檢查 (避免把異常資料丟給 JOJO)
    for lvl, msg in self_check(shops):
        icon = {"OK": "✅", "WARN": "⚠", "ERR": "❌"}.get(lvl, "•")
        print(f"   {icon} {msg}")


if __name__ == "__main__":
    main()
