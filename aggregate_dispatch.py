"""aggregate_dispatch.py — ANN(客服業助) 把 JOJO 配車完的產出，按日期收進
桌面/客服AI/<日期>/，並彙整成一份 Excel（路線報表 + 地圖表）。

用途：每天 JOJO 跑完路線後，ANN 把資料歸檔 + 整理成好看的單一 Excel。
  - 路線報表 工作表：來自 route_report.csv（每站 車號/序號/店家/地址/瓶數/品項/
    下貨秒數/預計到店/預計離店 + 路線總計區塊）
  - 地圖表 工作表：嵌入 route_map.png（若有分車 png 也一併嵌入）

用法：
  python aggregate_dispatch.py [日期]            # 日期預設=今天(YYYY-MM-DD)
  python aggregate_dispatch.py 2026-07-18
  python aggregate_dispatch.py --src <來源資料夾> --dest <目標資料夾>

來源預設：桌面/當日車輛報表/<日期>/
目標預設：桌面/客服AI/<日期>/
"""
import os
import shutil
import csv
import argparse
import datetime

try:
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    Workbook = XLImage = None

ONEDRIVE_DESKTOP = os.path.join(os.path.expanduser("~"), "OneDrive", "桌面")
REPORT_DIR = os.path.join(ONEDRIVE_DESKTOP, "當日車輛報表")
DEST_ROOT = os.path.join(ONEDRIVE_DESKTOP, "客服AI")

_STATION_HEADER = ["車號", "序號", "店家", "地址", "瓶數", "品項", "下貨秒數", "預計到店", "預計離店"]


def _parse_report_csv(csv_path):
    """回傳 (stations:list[list], totals:list[list])。 stations=逐站資料列；totals=總計區塊。"""
    stations, totals = [], []
    if not os.path.exists(csv_path):
        return stations, totals
    with open(csv_path, encoding="utf-8-sig", newline="") as fh:
        rows = list(csv.reader(fh))
    mode = "stations"
    for row in rows:
        if not row or all(str(c).strip() == "" for c in row):
            if stations:
                mode = "totals"
            continue
        if str(row[0]).startswith("==="):
            mode = "totals"
            continue
        if row[0] == "車號" and "店家" in row:  # 表頭列，跳過
            continue
        if mode == "stations":
            stations.append(row)
        else:
            totals.append(row)
    return stations, totals


def _style_header(ws, row, ncols):
    fill = PatternFill("solid", fgColor="0B6B3A")
    font = Font(bold=True, color="FFFFFF", size=11)
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _build_excel(dest, date_str, stations, totals, map_files):
    if Workbook is None:
        raise RuntimeError("請先安裝 openpyxl: uv pip install openpyxl")
    wb = Workbook()

    # ---- 工作表1：路線報表 ----
    ws = wb.active
    ws.title = "路線報表"
    ws.append([f"每日配送彙整 — {date_str}"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(_STATION_HEADER))
    ws.cell(row=1, column=1).font = Font(bold=True, size=14, color="0B6B3A")
    ws.append(_STATION_HEADER)
    _style_header(ws, 2, len(_STATION_HEADER))
    for r in stations:
        # 補齊 9 欄（品項可能含逗號已被 csv 正確解析）
        padded = (r + [""] * len(_STATION_HEADER))[:len(_STATION_HEADER)]
        ws.append(padded)
    # 總計區塊
    if totals:
        ws.append([])
        ws.append(["路線總計"])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12, color="0B6B3A")
        for t in totals:
            ws.append(t)
    ws.freeze_panes = "A3"
    # 欄寬
    widths = [8, 6, 30, 38, 7, 40, 9, 9, 9]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    # ---- 工作表2：地圖表（嵌入 PNG） ----
    ws2 = wb.create_sheet("地圖表")
    ws2.append([f"配送路線地圖 — {date_str}"])
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1)
    ws2.cell(row=1, column=1).font = Font(bold=True, size=14, color="0B6B3A")
    row_cursor = 3
    for png in map_files:
        if not os.path.exists(png):
            continue
        try:
            img = XLImage(png)
            MAX_W = 1150
            if img.width > MAX_W:
                ratio = MAX_W / img.width
                img.width = int(img.width * ratio)
                img.height = int(img.height * ratio)
            ws2.add_image(img, f"A{row_cursor}")
            row_cursor += int(img.height / 18) + 4  # 粗略換行
        except Exception as e:
            ws2.cell(row=row_cursor, column=1, value=f"（嵌入地圖失敗：{png} -> {e}）")
            row_cursor += 2

    out_path = os.path.join(dest, f"每日配送彙整_{date_str}.xlsx")
    wb.save(out_path)
    return out_path


def aggregate(date_str=None, src=None, dest=None):
    if date_str is None:
        date_str = datetime.datetime.now().strftime("%Y-%m-%d")
    src = src or os.path.join(REPORT_DIR, date_str)
    dest = dest or os.path.join(DEST_ROOT, date_str)
    if not os.path.isdir(src):
        raise FileNotFoundError(f"找不到 JOJO 報表資料夾：{src}\n（JOJO 還沒配車完？或日期給錯？）")

    os.makedirs(dest, exist_ok=True)

    # 1) 複製原始報表檔到目標資料夾（html/csv/png 全部保留）
    copied = []
    for f in sorted(os.listdir(src)):
        s = os.path.join(src, f)
        if os.path.isfile(s):
            shutil.copy2(s, os.path.join(dest, f))
            copied.append(f)

    # 2) 解析路線報表
    csv_path = os.path.join(src, "route_report.csv")
    stations, totals = _parse_report_csv(csv_path)

    # 3) 找地圖 PNG（合併圖 + 分車圖）
    map_files = []
    for f in sorted(os.listdir(src)):
        if f.lower().endswith(".png"):
            map_files.append(os.path.join(src, f))

    # 4) 建 Excel
    xlsx_path = _build_excel(dest, date_str, stations, totals, map_files)

    return {
        "date": date_str,
        "src": src,
        "dest": dest,
        "copied": copied,
        "xlsx": xlsx_path,
        "n_stations": len(stations),
        "n_maps": len(map_files),
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("date", nargs="?", default=None, help="日期 YYYY-MM-DD（預設今天）")
    ap.add_argument("--src", default=None, help="來源報表資料夾（覆寫預設）")
    ap.add_argument("--dest", default=None, help="目標資料夾（覆寫預設）")
    args = ap.parse_args()
    try:
        r = aggregate(args.date, args.src, args.dest)
    except FileNotFoundError as e:
        print(f"❌ {e}")
        return
    print(f"✅ 彙整完成：{r['date']}")
    print(f"   來源：{r['src']}")
    print(f"   歸檔至：{r['dest']}")
    print(f"   複製檔案：{len(r['copied'])} 個 -> {', '.join(r['copied'])}")
    print(f"   路線報表站數：{r['n_stations']} ｜ 地圖圖檔：{r['n_maps']}")
    print(f"   Excel：{r['xlsx']}")


if __name__ == "__main__":
    main()
